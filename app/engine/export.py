from __future__ import annotations

import csv
import io

from openpyxl import Workbook, load_workbook


def _neutralize(s: str) -> str:
    # CSV/spreadsheet formula-injection guard: a leading =,+,-,@ is treated as a
    # formula by Excel/Sheets, so prefix it with a single quote to force text.
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _stringify(value) -> str:
    if value is None:
        out = ""
    elif isinstance(value, (list, tuple)):
        out = "; ".join(str(v) for v in value)
    elif isinstance(value, dict):
        out = "; ".join(f"{k}={v}" for k, v in value.items())
    elif isinstance(value, bool):
        out = "Y" if value else "N"
    else:
        out = str(value)
    return _neutralize(out)


def rows_to_csv(columns: list[str], rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_stringify(row.get(c, "")) for c in columns])
    return buf.getvalue().encode("utf-8")


def rows_to_xlsx(sheet_name: str, columns: list[str], rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel tab name limit
    ws.append(columns)
    for row in rows:
        ws.append([_stringify(row.get(c, "")) for c in columns])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def append_xlsx(existing: bytes, sheet_name: str, rows: list[dict]) -> bytes:
    wb = load_workbook(io.BytesIO(existing))
    name = sheet_name[:31]
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    header = [c.value for c in ws[1] if c.value is not None]
    if not header:
        header = list(rows[0].keys()) if rows else []
        ws.append(header)
    for row in rows:
        ws.append([_stringify(row.get(col, "")) for col in header])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
