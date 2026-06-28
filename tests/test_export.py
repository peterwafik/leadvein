import io
from openpyxl import load_workbook

from app.engine.export import rows_to_csv, rows_to_xlsx, append_xlsx

COLUMNS = ["name", "website", "email"]
ROWS = [
    {"name": "Mario's", "website": "marios.com", "email": "info@marios.com"},
    {"name": "Joe's", "website": "joes.com", "email": ""},
]


def test_rows_to_csv_has_header_and_rows():
    data = rows_to_csv(COLUMNS, ROWS).decode("utf-8")
    lines = [l for l in data.splitlines() if l]
    assert lines[0] == "name,website,email"
    assert "Mario's,marios.com,info@marios.com" in lines[1]
    assert len(lines) == 3


def test_rows_to_xlsx_named_sheet_and_content():
    blob = rows_to_xlsx("GloriaFood Prospects", COLUMNS, ROWS)
    wb = load_workbook(io.BytesIO(blob))
    assert "GloriaFood Prospects" in wb.sheetnames
    ws = wb["GloriaFood Prospects"]
    assert [c.value for c in ws[1]] == COLUMNS
    assert ws.cell(row=2, column=1).value == "Mario's"


def test_append_xlsx_preserves_existing_column_order():
    existing_cols = ["website", "name", "email", "status"]
    existing = rows_to_xlsx("GloriaFood Prospects", existing_cols,
                            [{"website": "old.com", "name": "Old", "email": "",
                              "status": "Contacted"}])
    merged = append_xlsx(existing, "GloriaFood Prospects", ROWS)
    wb = load_workbook(io.BytesIO(merged))
    ws = wb["GloriaFood Prospects"]
    assert [c.value for c in ws[1]] == existing_cols  # order preserved
    assert ws.cell(row=2, column=1).value == "old.com"
    assert ws.cell(row=3, column=2).value == "Mario's"  # name is col 2 here


def test_csv_neutralizes_formula_injection():
    data = rows_to_csv(["name", "note"],
                       [{"name": "=cmd", "note": "+1"},
                        {"name": "-2+3", "note": "@x"}]).decode("utf-8")
    # leading =,+,-,@ are prefixed with a single quote so spreadsheets render as text
    assert "'=cmd" in data
    assert "'+1" in data
    assert "'-2+3" in data
    assert "'@x" in data
    # a normal value is untouched
    safe = rows_to_csv(["name"], [{"name": "Mario's Diner"}]).decode("utf-8")
    assert "Mario's Diner" in safe and "'Mario" not in safe
