"""Admin bulk unlock/export: bypasses the ECONOMY, never the COMPLIANCE spine."""
from __future__ import annotations

import json
import re

from fastapi.testclient import TestClient
from sqlmodel import Session, select

import app.leadvault as lv
from app.core.db import CreditTransaction, Lead, PurchasedLead
from app.quality.ordinals import apply_tier_columns


def _ensure_admin():
    """Create admin@demo.local if not already present."""
    from app.core.db import User
    from app.core.auth import create_user
    with Session(lv.engine) as s:
        if not s.exec(select(User).where(User.email == "admin@demo.local")).first():
            create_user(s, "admin@demo.local", "admin12345", role="admin")


def _login(c, email, pw):
    m = re.search(r'name="csrf_token" value="([^"]+)"', c.get("/login").text)
    token = m.group(1)
    c.post("/login", data={"email": email, "password": pw, "csrf_token": token},
           follow_redirects=False)
    return token


def _seed_lead(s, name="Bulk Exportable"):
    val = {"profile": {"tier": "validated"}, "phone": {"tier": "validated"},
           "email": {"tier": "present"}, "address": {"tier": "present"},
           "website": {"tier": "absent"}}
    lead = Lead(business_name=name, city="Exportville", country="GB",
                phone="+441865222222", score_total=80,
                category_keys_json='["bakery"]',
                validation_json=json.dumps(val),
                attribution="© OpenStreetMap contributors (ODbL)",
                retention_expiry="2999-01-01T00:00:00+00:00")
    apply_tier_columns(lead, val)
    s.add(lead); s.commit(); s.refresh(lead)
    return lead.id


def test_reveal_returns_detail_without_economy_side_effects():
    _ensure_admin()
    with Session(lv.engine) as s:
        lid = _seed_lead(s)
        n_purch = len(s.exec(select(PurchasedLead)).all())
        n_tx = len(s.exec(select(CreditTransaction)).all())
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/reveal", json={"lead_ids": [lid]},
               headers={"X-CSRF-Token": token})
    assert r.status_code == 200
    row = r.json()["leads"][0]
    assert row["phone"] == "+441865222222"          # full detail for admin
    with Session(lv.engine) as s:
        assert len(s.exec(select(PurchasedLead)).all()) == n_purch
        assert len(s.exec(select(CreditTransaction)).all()) == n_tx
        assert s.get(Lead, lid).times_sold == 0


def test_reveal_skips_non_int_ids():
    """Mixed lead_ids [validId, 'abc', [1]] → 200 with only the valid lead returned."""
    _ensure_admin()
    with Session(lv.engine) as s:
        lid = _seed_lead(s, "Mixed ID Lead")
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/reveal", json={"lead_ids": [lid, "abc", [1]]},
               headers={"X-CSRF-Token": token})
    assert r.status_code == 200
    leads = r.json()["leads"]
    assert len(leads) == 1
    assert leads[0]["business_name"] == "Mixed ID Lead"


def test_export_csv_full_detail_and_attribution():
    _ensure_admin()
    with Session(lv.engine) as s:
        lid = _seed_lead(s, "Bulk CSV Lead")
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/export",
               data={"csrf_token": token, "format": "csv",
                     "lead_ids": str(lid)})
    assert r.status_code == 200
    body = r.content.decode("utf-8", errors="replace")
    assert "Bulk CSV Lead" in body and "+441865222222" in body
    assert "OpenStreetMap contributors" in body      # ODbL attribution embedded
    assert "validated" in body                       # tier labels exported


def test_export_xlsx_roundtrip():
    _ensure_admin()
    with Session(lv.engine) as s:
        lid = _seed_lead(s, "Bulk XLSX Lead")
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/export",
               data={"csrf_token": token, "format": "xlsx", "lead_ids": str(lid)})
    assert r.status_code == 200
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    text = " ".join(str(c.value) for row in wb.active.iter_rows() for c in row)
    assert "Bulk XLSX Lead" in text


def test_buyer_cannot_touch_bulk_endpoints_or_see_controls():
    c = TestClient(lv.app)
    token = _login(c, "buyer@demo.local", "buyer12345")
    r = c.post("/admin/bulk/reveal", json={"lead_ids": [1]},
               headers={"X-CSRF-Token": token})
    assert r.status_code in (302, 303, 401, 403)
    html = c.get("/app/find").text
    assert "Unlock selected" not in html
    assert "Export all" not in html


def test_admin_can_view_find_page_with_controls():
    _ensure_admin()
    c = TestClient(lv.app)
    _login(c, "admin@demo.local", "admin12345")
    r = c.get("/app/find")
    assert r.status_code == 200
    assert "Unlock selected" in r.text and "Export all" in r.text


def _scripts(html: str) -> str:
    """The <script> portions only — proves the JS is WIRED, not just markup present."""
    return "\n".join(re.findall(r"<script[^>]*>(.*?)</script>", html, flags=re.S))


def test_admin_find_script_wires_bulk_js():
    """The find page JS must actually reference the bulk endpoints (dead-markup guard)."""
    _ensure_admin()
    c = TestClient(lv.app)
    _login(c, "admin@demo.local", "admin12345")
    script = _scripts(c.get("/app/find").text)
    assert "adminExportComposition" in script      # export-all composition binding wired
    assert "/admin/bulk/reveal" in script          # unlock/reveal fetch wired


def test_buyer_find_has_no_bulk_js():
    """Buyer HTML must contain neither the reveal URL nor the composition binding."""
    c = TestClient(lv.app)
    _login(c, "buyer@demo.local", "buyer12345")
    html = c.get("/app/find").text
    assert "adminExportComposition" not in html
    assert "/admin/bulk/reveal" not in html


def test_export_is_uncapped():
    """Operator decision 2026-07-04: exports have NO row cap — the owner may dump
    the full serveable inventory in one file. This replaces the old 10k-cap test."""
    _ensure_admin()
    ids = []
    with Session(lv.engine) as s:
        for i in range(3):
            ids.append(_seed_lead(s, f"Uncap Row {i}"))
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/export",
               data={"csrf_token": token, "format": "csv",
                     "lead_ids": ",".join(str(i) for i in ids)})
    assert r.status_code == 200
    body = r.content.decode("utf-8", errors="replace")
    for i in range(3):
        assert f"Uncap Row {i}" in body


def test_global_suppression_hides_lead_from_reveal_and_export():
    """Owner path applies GLOBAL suppression (matches the on-screen estimate)."""
    from app.core.db import SuppressionEntry, SuppressionList
    _ensure_admin()
    # Unique phone so we don't suppress the shared-phone leads other tests seed.
    uniq_phone = "+441111000999"
    val = {"profile": {"tier": "present"}, "phone": {"tier": "validated"},
           "email": {"tier": "absent"}, "address": {"tier": "absent"},
           "website": {"tier": "absent"}}
    with Session(lv.engine) as s:
        lead = Lead(business_name="Globally Suppressed Co", city="Nowhere", country="GB",
                    phone=uniq_phone, score_total=70, category_keys_json='["bakery"]',
                    validation_json=json.dumps(val),
                    attribution="© OpenStreetMap contributors (ODbL)",
                    retention_expiry="2999-01-01T00:00:00+00:00")
        apply_tier_columns(lead, val)
        s.add(lead); s.commit(); s.refresh(lead)
        lid = lead.id
        sl = SuppressionList(buyer_account_id=None, name="global test list")  # None => global
        s.add(sl); s.commit(); s.refresh(sl)
        s.add(SuppressionEntry(list_id=sl.id, kind="phone", value=uniq_phone))
        s.commit()
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")
    r = c.post("/admin/bulk/reveal", json={"lead_ids": [lid]},
               headers={"X-CSRF-Token": token})
    assert r.status_code == 200
    assert r.json()["leads"] == []                 # globally suppressed → absent from reveal
    r2 = c.post("/admin/bulk/export",
                data={"csrf_token": token, "format": "csv", "lead_ids": str(lid)})
    assert r2.status_code == 200
    assert "Globally Suppressed Co" not in r2.content.decode("utf-8", "replace")


def test_export_degenerate_composition_returns_400():
    """Degenerate compositions must be rejected with 400, not 500 (KeyError guard)."""
    import json as _json
    _ensure_admin()
    c = TestClient(lv.app)
    token = _login(c, "admin@demo.local", "admin12345")

    # {} — empty dict, no "op" and no "predicate"
    r = c.post("/admin/bulk/export",
               data={"csrf_token": token, "format": "csv",
                     "composition": _json.dumps({})})
    assert r.status_code == 400
    assert "Invalid targeting composition" in r.text

    # Node missing "predicate" — would KeyError inside matching_by_composition
    bad_comp = {"op": "AND", "nodes": [{"params": {}}]}
    r2 = c.post("/admin/bulk/export",
                data={"csrf_token": token, "format": "csv",
                      "composition": _json.dumps(bad_comp)})
    assert r2.status_code == 400
    assert "Invalid targeting composition" in r2.text
