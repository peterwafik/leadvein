#!/usr/bin/env python3
"""
gloriafood_finder.py — find restaurants running GloriaFood and pull public contact info.

Two modes:
  discover : ask PublicWWW for sites embedding the GloriaFood widget (needs a free/paid
             PublicWWW API key -> https://publicwww.com/  env: PUBLICWWW_KEY)
  verify   : for a list of domains, fetch each homepage, confirm GloriaFood, extract contacts

Output: appends rows into the 'GloriaFood Prospects' tab of an .xlsx (and a .csv mirror).

Examples
  # 1) discover domains via PublicWWW (writes domains.txt)
  PUBLICWWW_KEY=xxxx python gloriafood_finder.py discover --limit 500 --out domains.txt

  # 2) verify + enrich a domain list into the tracker
  python gloriafood_finder.py verify --in domains.txt \
        --xlsx GloriaFood_Outreach_Tracker.xlsx --csv prospects.csv

  # verify also accepts a single site:  python gloriafood_finder.py verify --url someplace.com
"""
import argparse, csv, os, re, sys, time, json
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UA = "Mozilla/5.0 (compatible; GloriaFoodResearch/1.0; +contact your own outreach)"
HEADERS = {"User-Agent": UA, "Accept-Language": "en"}

# --- GloriaFood fingerprints -------------------------------------------------
FINGERPRINTS = [
    "fbgcdn.com",          # GloriaFood / FoodBooking CDN that serves the widget
    "ewm2.js",             # the "easy web menu" embedder script
    "data-glf-cuid",       # customer id attribute on the ordering button
    "data-glf-ruid",       # restaurant id attribute on the ordering button
    "gloriafood",          # generic mentions / menu.gloriafood.com
]
RUID_RE  = re.compile(r'data-glf-ruid=["\']([0-9a-fA-F-]+)["\']')
CUID_RE  = re.compile(r'data-glf-cuid=["\']([0-9a-fA-F-]+)["\']')
EMAIL_RE = re.compile(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}')
PHONE_RE = re.compile(r'(\+?\d[\d\s().-]{7,}\d)')

PROSPECT_TAB = "GloriaFood Prospects"


def norm_url(d):
    d = d.strip()
    if not d:
        return None
    if not d.startswith(("http://", "https://")):
        d = "https://" + d
    return d


def fetch(url, timeout=15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and r.text:
            return r.text
    except requests.RequestException:
        return None
    return None


def analyse(html, url):
    low = html.lower()
    hit = next((f for f in FINGERPRINTS if f in low), None)
    soup = BeautifulSoup(html, "html.parser")

    name = ""
    if soup.title and soup.title.string:
        name = soup.title.string.strip()
    ogn = soup.find("meta", property="og:site_name") or soup.find("meta", property="og:title")
    if ogn and ogn.get("content"):
        name = ogn["content"].strip() or name

    emails = []
    for a in soup.select('a[href^="mailto:"]'):
        emails.append(a["href"].split("mailto:", 1)[1].split("?")[0])
    emails += EMAIL_RE.findall(html)
    emails = [e for e in dict.fromkeys(e.lower() for e in emails)
              if not e.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"))
              and "fbgcdn" not in e and "sentry" not in e and "wixpress" not in e]

    phones = [a["href"].split("tel:", 1)[1] for a in soup.select('a[href^="tel:"]')]
    if not phones:
        phones = PHONE_RE.findall(soup.get_text(" "))
    phones = list(dict.fromkeys(p.strip() for p in phones))

    ruid = (RUID_RE.search(html) or [None, ""])[1] if RUID_RE.search(html) else ""
    cuid = (CUID_RE.search(html) or [None, ""])[1] if CUID_RE.search(html) else ""

    return {
        "name": name,
        "website": url,
        "on_gf": "Y" if hit else "N",
        "ruid": ruid,
        "email": emails[0] if emails else "",
        "emails_all": "; ".join(emails[:5]),
        "phone": phones[0] if phones else "",
        "matched": hit or "",
    }


def discover(args):
    if args.engine == "publicwww":
        return discover_publicwww(args)
    return discover_urlscan(args)


def _save_domains(domains, out, engine):
    domains = list(dict.fromkeys(domains))[:None]
    with open(out, "w") as f:
        f.write("\n".join(domains))
    print(f"discover[{engine}]: {len(domains)} unique domains -> {out}")


def discover_urlscan(args):
    """FREE engine. urlscan.io indexes every domain a scanned page contacts,
    so pages that loaded the GloriaFood widget show up under domain:fbgcdn.com.
    Optional free API key (env URLSCAN_KEY) raises the rate limit."""
    query = args.query or "domain:fbgcdn.com"
    headers = dict(HEADERS)
    key = os.environ.get("URLSCAN_KEY")
    if key:
        headers["API-Key"] = key
    domains, search_after = [], None
    while len(domains) < args.limit:
        url = ("https://urlscan.io/api/v1/search/?q=" +
               requests.utils.quote(query) + "&size=100")
        if search_after:
            url += f"&search_after={search_after}"
        try:
            r = requests.get(url, headers=headers, timeout=40)
            if r.status_code == 429:
                print("urlscan rate-limited; sleeping 60s (add URLSCAN_KEY to raise the limit)")
                time.sleep(60); continue
            data = r.json()
        except Exception as e:
            print(f"urlscan error: {e}"); break
        results = data.get("results", [])
        if not results:
            break
        for res in results:
            host = (res.get("page", {}).get("domain") or "").lower().replace("www.", "")
            # skip GloriaFood's own infra domains; we want the restaurants
            if host and not any(s in host for s in ("gloriafood", "fbgcdn", "foodbooking")):
                domains.append(host)
        if not data.get("has_more"):
            break
        search_after = ",".join(str(x) for x in results[-1].get("sort", []))
        if not search_after:
            break
        time.sleep(2)
    _save_domains(domains, args.out, "urlscan")


def discover_publicwww(args):
    key = os.environ.get("PUBLICWWW_KEY")
    if not key:
        sys.exit("Set PUBLICWWW_KEY (https://publicwww.com/). "
                 "Or use the free default: --engine urlscan. "
                 "Or skip discover and feed your own domains.txt to 'verify'.")
    query = args.query or '"fbgcdn.com/embedder"'
    domains, page = [], 1
    while len(domains) < args.limit:
        url = ("https://publicwww.com/websites/" +
               requests.utils.quote(query) +
               f"/?export=urls&page={page}&key={key}")
        txt = fetch(url, timeout=40)
        if not txt:
            break
        rows = [l.strip() for l in txt.splitlines() if l.strip() and "." in l]
        if not rows:
            break
        for l in rows:
            host = urlparse(norm_url(l)).netloc.lower().replace("www.", "")
            if host:
                domains.append(host)
        page += 1
        time.sleep(1)
    _save_domains(domains, args.out, "publicwww")


def load_domains(args):
    if args.url:
        return [args.url]
    with open(args.infile) as f:
        return [l.strip() for l in f if l.strip() and not l.startswith("#")]


def append_xlsx(path, rows):
    wb = load_workbook(path)
    ws = wb[PROSPECT_TAB] if PROSPECT_TAB in wb.sheetnames else wb.create_sheet(PROSPECT_TAB)
    start = ws.max_row + 1
    # write into the existing column order of the tracker
    for i, r in enumerate(rows, start=start):
        vals = [r["name"], r["website"], "", r["email"], r["phone"], r["on_gf"],
                r["ruid"], "", "", "", "PublicWWW", "", "Not contacted", "", r["emails_all"]]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=i, column=c, value=v)
    wb.save(path)


def verify(args):
    domains = load_domains(args)
    out_rows, seen = [], set()
    for i, d in enumerate(domains, 1):
        url = norm_url(d)
        host = urlparse(url).netloc.lower()
        if host in seen:
            continue
        seen.add(host)
        html = fetch(url)
        if not html and not url.endswith("/"):
            html = fetch(url + "/")
        if not html:
            print(f"[{i}/{len(domains)}] {host}: no response")
            continue
        info = analyse(html, url)
        out_rows.append(info)
        print(f"[{i}/{len(domains)}] {host}: GloriaFood={info['on_gf']} "
              f"name={info['name'][:40]!r} email={info['email']} phone={info['phone']}")
        time.sleep(args.delay)

    gf = [r for r in out_rows if r["on_gf"] == "Y"]
    print(f"\n{len(out_rows)} checked, {len(gf)} confirmed on GloriaFood.")

    if args.csv:
        with open(args.csv, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()) if out_rows else
                               ["name", "website", "on_gf", "ruid", "email", "emails_all", "phone", "matched"])
            w.writeheader(); w.writerows(out_rows)
        print(f"csv  -> {args.csv}")
    if args.xlsx:
        append_xlsx(args.xlsx, gf if args.only_gf else out_rows)
        print(f"xlsx -> {args.xlsx}  (appended {len(gf if args.only_gf else out_rows)} rows to '{PROSPECT_TAB}')")


def main():
    ap = argparse.ArgumentParser(description="Find restaurants on GloriaFood and pull public contacts.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    d = sub.add_parser("discover", help="get candidate domains (free urlscan.io by default)")
    d.add_argument("--engine", choices=["urlscan", "publicwww"], default="urlscan",
                   help="urlscan.io is free (default); publicwww needs a paid-ish key")
    d.add_argument("--query", default=None,
                   help='search query (urlscan default: domain:fbgcdn.com)')
    d.add_argument("--limit", type=int, default=500)
    d.add_argument("--out", default="domains.txt")
    d.set_defaults(func=discover)

    v = sub.add_parser("verify", help="confirm GloriaFood + extract contacts for a domain list")
    v.add_argument("--in", dest="infile", default="domains.txt")
    v.add_argument("--url", default=None, help="check a single site instead of a file")
    v.add_argument("--xlsx", default=None, help="tracker .xlsx to append into")
    v.add_argument("--csv", default=None, help="also write a CSV")
    v.add_argument("--delay", type=float, default=1.0, help="seconds between requests (be polite)")
    v.add_argument("--only-gf", action="store_true", help="only write confirmed-GloriaFood rows to xlsx")
    v.set_defaults(func=verify, only_gf=True)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
